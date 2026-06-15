"""Write the formatted Excel ops report."""
import os

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


class ExcelReport:
    """Build a multi-sheet Excel report with simple formatting."""

    def __init__(self, path: str):
        self.path = path
        os.makedirs(os.path.dirname(path) or ".", exist_ok=True)

    def write(self, sheets: dict[str, pd.DataFrame]) -> None:
        with pd.ExcelWriter(self.path, engine="openpyxl") as writer:
            for sheet_name, df in sheets.items():
                safe_name = sheet_name[:31]
                df.to_excel(writer, sheet_name=safe_name, index=False)
                ws = writer.sheets[safe_name]

                header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
                header_font = Font(color="FFFFFF", bold=True)
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

                for idx in range(1, ws.max_column + 1):
                    col = get_column_letter(idx)
                    max_len = max(
                        (len(str(cell.value)) for cell in ws[col] if cell.value is not None),
                        default=0,
                    )
                    ws.column_dimensions[col].width = min(max_len + 2, 50)

                ws.freeze_panes = "A2"

        print(f"Report saved: {self.path}")
